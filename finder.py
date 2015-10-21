import os
from openpyxl import load_workbook
from openpyxl import Workbook


wb = load_workbook(filename='C:\\Users\\patrizio\\Documents\\GitHub\\minixl\\test_data\\t.xlsx',use_iterators=True)

def get_company_names():
	ws = wb['Sheet1']
	result = []
	for row in ws.iter_rows(row_offset=1):
		if row[1].value:
			company = row[1].value.strip()
			result.append(company)
	return result
		
		

def hash_year_values():
	ws=wb['Sheet2']
	result = {}
	for row in ws.iter_rows(range_string = 'D1:CU1'):
		for cell in row:
			if isinstance(cell.value,int):			
				result[cell.column] = cell.value
	return result
	
def hash_event_years():
	ws=wb['Sheet2']
	result={}
	company_list = get_company_names()
	year_dict = hash_year_values()
	for row in ws.iter_rows(row_offset=1):
		company_name = row[0].value.strip()
		if company_name in company_list:	 
			for cell in row:
				if cell.column in year_dict and cell.value:
					result[company_name] = year_dict[cell.column]
					break
	return result 

	
def write_event_years():
	wb = load_workbook(filename='C:\\Users\\patrizio\\Documents\\GitHub\\minixl\\test_data\\t.xlsx')
	ws=wb['Sheet1']
	event_year_dict = hash_event_years()
	companies = []
	prevcompany = ''
	for row in range(2,ws.get_highest_row()):
		Cell = ws.cell(column=2,row=row)
		cell_value = unicode(Cell.value).strip()
		if cell_value and cell_value != prevcompany and cell_value in event_year_dict:
			company_name = cell_value
			ws.cell(column=3,row=row).value = event_year_dict[company_name]
			companies.append(company_name)
			prevcompany = cell_value
	wb.save('C:\\Users\\patrizio\\Documents\\GitHub\\minixl\\test_data\\t.xlsx')
	print "New values written to spreadsheet"
	return companies	
	
# def create_new_xl():
	# nb = Workbook(write_only=True)
	# ws = nb.create_sheet()
	# entry_list = hash_event_years()
	# for entry in entry_list:
		# ws.append([entry,entry_list[entry],entry_list[entry]-1])
	# nb.save('C:\\Users\\patrizio\\Documents\\GitHub\\minixl\\test_data\\event_years.xlsx')
	# print "Saved new workbook with eventyears"
	
# companies_sheet1 = get_company_names()
# companies_sheet2 = hash_event_years()
# for i in companies_sheet1:
	# if i not in companies_sheet2:
		# print i 
# print len(get_company_names())						
#print get_event_year()		
#print hash_year_columns()
write_event_years()
#create_new_xl()		