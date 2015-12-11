from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime
from os.path import basename
import itertools
import string
import os
import csv

# old_doc = "C:\\Users\\Alec\\.projects\\minixl\\test_data\\t.xlsx"
old_doc = "C:\\Users\\patrizio\\Documents\\GitHub\\minixl\\test_data\\t.xlsx"
# target_firm_doc = "C:\\Users\\Alec\\.projects\\minixl\\test_data\\3_Target_firm.xlsx"
target_firm_doc = "C:\\Users\patrizio\\Documents\\GitHub\\minixl\\test_data\\t.xlsx"
industry_firms_doc = "C:\\Users\\Alec\\.projects\\minixl\\test_data\\2_BigData.xlsx"
output_path = "C:\\Users\\patrizio\\Documents\\GitHub\\minixl\\test_data"
#set the range string which contains the range of header data to be used in hash_year_values
header_cell_range = 'D1:CU1'
wb = load_workbook(filename=target_firm_doc,use_iterators=True)
sheets = wb.get_sheet_names()
target_firm_sheet = sheets[0]
# wb2 = load_workbook(filename=industry_firms_doc,use_iterators=True)

wb3 = load_workbook(filename=old_doc,use_iterators=True)
sheets_old = wb3.get_sheet_names()
rank_sheet = sheets_old[1]

#decorator to convert indexes to 0-based
def convertindex(func):
	def minus_one(index_string):
		return func(index_string) - 1
	return minus_one
get_index = convertindex(column_index_from_string)	
	
def replace_punc_with(text):
	"""
	Removes all punctuation from a string.
	"""
	exclude = set(string.punctuation)
	return ''.join(ch for ch in text if ch not in exclude)

def word_similar(word1,word2):
	"""
	Evaluates the similarity of two words, based on the number and order of characters.
	Returns True if words are > 80% similar or share a common word not in the list of exlusion suffixes.
	"""
	word1 = replace_punc_with(word1.lower().strip())
	word2 = replace_punc_with(word2.lower().strip())
	count = 0
	for char in word1:
		if char in word2 and word1.index(char) == word2.index(char):
			count += 1
		if 0.8 <= count / len(word1) <= 1:
			return True
	for word in word1.split():
		if word in word2.split() and word not in ["inc","co","ltd","corp","cl","cp"]:
			return True
	return False





def get_company_names():
	ws = wb[target_firm_sheet]
	result = []
	for row in ws.iter_rows(row_offset=1):
		if row[1].value:
			company = unicode(row[1].value).strip()
			result.append(company)
	return result



def hash_year_values():
	ws=wb3[rank_sheet]
	result = {}
	for row in ws.iter_rows(range_string = header_cell_range):
		for cell in row:
			value = cell.value
			if isinstance(value,long) or isinstance(value,int):
				result[cell.column] = value
	return result

def hash_event_years(company_names,year_col_letter_dict):
	ws=wb3[rank_sheet]
	result={}
	company_list = get_company_names()
	year_dict = hash_year_values()
	for row in ws.iter_rows(row_offset=1):
		company_name = unicode(row[0].value).strip()
		event_years =[]
		if company_name in company_list:
			for cell in row:
				if cell.column in year_dict and cell.value:
					event_years.append(year_dict[cell.column])
			event_years.sort()
			result[company_name] = event_years
	return result

def get_ranks(event_year_dict,year_col_letter_dict):
	ws = wb3[rank_sheet]
	result = {}
	event_years = event_year_dict
	year_column_dict = year_col_letter_dict
	for row in ws.iter_rows(row_offset=1):
		company_name = unicode(row[0].value).strip()
		if company_name in event_years:
			result[company_name] = {}
			for cell in row:
				if cell.column in year_column_dict:
					year = year_column_dict[cell.column]
					if year in event_years[company_name]:
						result[company_name][year] = cell.value
	return result



def write_event_years():
	wb = load_workbook(filename=target_firm_doc)
	ws=wb[sheets[0]]
	event_year_dict = hash_event_years()
	companies = []
	prevcompany = ''
	for row in range(2,ws.max_row):
		Cell = ws.cell(column=2,row=row)
		company = unicode(Cell.value).strip()
		if company and company != prevcompany and company in event_year_dict:
			ws.cell(column=3,row=row).value = event_year_dict[company][0]
			companies.append(company)
			prevcompany = company
	wb.save(target_firm_doc)
	print "New values written to spreadsheet"
	return companies

def check_pre_event_year():
	'''To find the pre-event year, this function checks
	the 1st event year, minus one, against the values in the
	Data Date column (first 4 digits). The program will iterate up from
	the starting row until it finds a match, or it reaches the end of the
	firms entries. If no match is found, the program will use
	the next event year in the hash_event_years dict, to calculate
	a new pre-event year, and iterate through the Data Dates again. This process repeats
	until there is a match, or the event year list is exausted. If the list is exausted,
	the firm name is appended to a dict (no_pre_event year).
	Once it finds a match, the program checks that same row,column=Net Income(loss)
	for a value. If there is no value, append the firm name to a dict (no_net_income_data) with firm name:pre-event year.
	'''
	ws=wb[target_firm_sheet]
	event_years = hash_event_years()
	net_income_col = get_index('AF')
	print "Net income col:" + str(net_income_col)
	no_net_income_data = []
	no_first_pre_event_year_data = []
	company_checked = {}
	for row in ws.iter_rows(row_offset=1):
		if row[1].value:
			firm = unicode(row[1].value).strip()
			pre_event_years = [int(x-1) for x in event_years[firm]]
			print firm + " has pre-event years:\n {0}\n".format(pre_event_years)
		if firm not in company_checked and row[13].value:
			datecell = row[13].value
			date = int(str(datecell)[:4])
			net_income = row[net_income_col].value
			if date in pre_event_years:
				if net_income:
					print firm + " has net income data for pre-event year {0}\n\n".format(date)
					company_checked[firm] = date
					if date != min(pre_event_years):
						no_first_pre_event_year_data.append(firm)
				elif firm not in no_net_income_data:
					no_net_income_data.append(firm)
	no_net_income_data = [x for x in no_net_income_data if x not in company_checked]
	no_pre_event_year = [e for e in event_years if e not in company_checked and e not in no_net_income_data]
	total_no_data_entries = no_net_income_data + no_pre_event_year
	changed_event_year = {i:company_checked[i] for i in no_first_pre_event_year_data}
	print "Firms that have pre-event year data, but no available net income data :\n" + str(no_net_income_data) + "\n\n" + "Firms with no available pre-event year data :\n"+ str(no_pre_event_year)+ "\n"
	with open("log.txt","w") as logfile:
		logfile.write("*" * 50 + "\n" + datetime.now().strftime('%H:%M %d/%m/%Y') + "\n\n" + "Log for analysis of: " + basename(target_firm_doc) + "\n" + "*" * 50 + "\n\n" + str(len(no_net_income_data)) + " Firms that have pre-event year data, but no available net income data :\n\n")
		for i in sorted(no_net_income_data):
			logfile.write(i + "\n")
		logfile.write("\n" + '*' * 50 + "\n" + str(len(no_pre_event_year)) + " Firms with no available pre-event year data :\n\n")
		for i in sorted(no_pre_event_year):
			logfile.write(i + "\n")
		logfile.write("\n" + "*" * 50 + "\n" + str(len(no_first_pre_event_year_data)) + " Firms for which net income data was not found for the first pre-event year , but was found for a following pre-event year: \n\n")
		for i in sorted(no_first_pre_event_year_data):
			logfile.write(i + "--" + str(company_checked[i]) + "\n")
		print "Results written to log.txt\n\n"
	return changed_event_year


def del_no_data_entries(entries_to_delete):
	''' (list) -> NoneType
	Iterates through the rows of spreadsheet and sets
	all cell values in the row to None if the value in column 2 (firm name)
	matches any value in entries_to_delete.
	Blank cells should be manually removed from the excel sheets
	after running this script.
	'''
	wb = load_workbook(filename=target_firm_doc)
	ws=wb[target_firm_sheet]
	for row in range(2,ws.max_row):
		Cell = ws.cell(column=2,row=row)
		if Cell.value:
			company = unicode(Cell.value).strip()
		if company in entries_to_delete:
			for col in range(2,ws.max_column + 1):
				cell = ws.cell(column=col,row=row)
				cell.value = None
	wb.save(target_firm_doc)
	return "All companies with no data have been removed from spreadsheet"

def change_entry(entries_to_change):
	''' (list) -> NoneType

	'''
	wb = load_workbook(filename=target_firm_doc)
	ws=wb[target_firm_sheet]
	for row in range(2,ws.max_row):
		Cell = ws.cell(column=2,row=row)
		if Cell.value:
			company = unicode(Cell.value).strip()
			pre_event_year = ws.cell(column=4,row=row)
			event_year = ws.cell(column=3,row=row)
			if company in entries_to_change:
				pre_event_year.value = int(entries_to_change[company])
			else:
				pre_event_year.value = event_year.value - 1
	wb.save(target_firm_doc)
	return "Changed event years"


def build_target_firm_data(rank_data):
	sic_col = get_index('BW')
	event_col = get_index('C')
	pre_event_col = get_index('D')
	company_col = get_index('U')
	name_col = get_index('B')
	assets_col = get_index('Y')
	net_income_col = get_index('AF')
	date_col = get_index('N')
	ranks = rank_data

	ws = wb[target_firm_sheet]
	result = {}
	for row in ws.iter_rows(row_offset=1):
		datecell = row[date_col].value
		total_assets = row[assets_col].value
		net_income = row[net_income_col].value
		if row[name_col].value:
			company = unicode(row[company_col].value).strip()
			name = unicode(row[name_col].value).strip()
			sic_code = row[sic_col].value
			eventyear = row[event_col].value
			pre_eventyear = row[pre_event_col].value
			result[company] = {"sic_code":sic_code,
				"eventyear":eventyear,
				"pre_eventyear":pre_eventyear,
				"total_assets":0,
				"net_income_pre_event_year":0,
				"net_income_event_year_plus":{},
				"100_best_ranks":ranks[name]
			}
		if datecell:
			date = int(str(datecell)[:4])
		if date == eventyear and total_assets:
			result[company]["total_assets"] = total_assets
		if date >= eventyear and net_income:
			result[company]["net_income_event_year_plus"][date] = net_income
		if date == pre_eventyear and net_income:
			result[company]["net_income_pre_event_year"] = net_income
	return result

def build_industry_groups(target_firm_data):
	ws = wb2.active
	target_firms = target_firm_data
	result = {}
	company_col = get_index('I')
	sic_col = get_index('BK')
	date_col = get_index('B')
	assets_col = get_index('M')
	net_income_col = get_index('T')
	for row in ws.iter_rows(row_offset=1):
		datecell = row[date_col].value
		date = int(str(datecell)[:4])
		sic_code = row[sic_col].value
		firm = unicode(row[company_col].value).strip()

		firm_total_assets = row[assets_col].value
		net_income = row[net_income_col].value
		if not firm_total_assets or not net_income:
			continue
		#TODO: need to fix name matching for edge cases
		target_firm_names = set([name for name in target_firms if target_firms[name]["sic_code"] == sic_code and name != firm])
		for name in target_firm_names:
			if date == target_firms[name]["eventyear"] and "total_assets" in target_firms[name] and 0.25 * target_firms[name]["total_assets"] <= firm_total_assets <= 2 * target_firms[name]["total_assets"]:
				if name in result:
					if not firm in result[name]:
						result[name][firm] = 0
				else:
					result[name] = {}
					result[name][firm] = 0
	no_total_assets = [i for i in target_firms if "total_assets" not in target_firms[i]]
	no_industry_size_match = [i for i in target_firms if i not in result and i not in no_total_assets]
	print "Target firms: {0}. Found matches for {1} target firms".format(len(target_firms),len(result))
	with open("no_industry_match_log.txt","w") as log:
		log.write("No event year data for these target firms (no total asset data):\n\n")
		for firm in no_total_assets:
			log.write(firm + "\n")
		log.write("\n\nNo size matched group for these companies: \n\n")
		for firm in no_industry_size_match:
			log.write(firm + "\n")

	return result

def get_income_data(industry_firm_data, target_firm_data):
	ws = wb2.active
	data = industry_firm_data
	target_firms = target_firm_data
	company_col = get_index('I')
	date_col = get_index('B')
	net_income_col = get_index('T') 
	for row in ws.iter_rows(row_offset=1):
		datecell = row[date_col].value
		date = int(str(datecell)[:4])
		firm = unicode(row[company_col].value).strip()

		net_income = row[net_income_col].value
		if not net_income:
			continue
		for target_firm in data:
			if firm in data[target_firm] and date == target_firms[target_firm]["pre_eventyear"]:
				data[target_firm][firm] = net_income
	return data



def get_match(income_data,target_firm_data):
	ws = wb2.active
	data = income_data
	target_firms = target_firm_data

	# company_col = get_index('I')
	# date_col = get_index('B')
	# net_income_col = get_index('T')
	for target_firm in data:
		target_firms[target_firm]["matched_firm"] = ''
		income_diffs = {}
		for match in data[target_firm]:
			income_diffs[match] = abs(target_firms[target_firm]["net_income_pre_event_year"] - data[target_firm][match])
		lowest_diff = min(income_diffs.values())
		for firm in income_diffs:
			if income_diffs[firm] == lowest_diff:
				target_firms[target_firm]["matched_firm"] = firm
	return target_firms





def create_match_output(output_file_name,entry_data,omitted_records = []):
	output = os.path.join(output_path,output_file_name)
	nb = Workbook(write_only=True)
	ws = nb.create_sheet()
	firms = [entry for entry in entry_data]
	sample = firms[0]
	headers = ["Target Firm"] + [record for record in entry_data[sample] if record not in omitted_records]
	ws.append(headers)
	for entry in entry_data:
		ws.append([entry] + [entry_data[entry][record] for record in entry_data[entry] if record in headers])
	nb.save(output)
	print "Saved new workbook to: {0}".format(output)
	

def create_year_output(output_file_name,entry_data,record, year_start=1998,year_end=2014):
	output = os.path.join(output_path,output_file_name)
	col_end = year_end - year_start + 1
	wb = Workbook()
	ws=wb.active
	firms = [entry for entry in entry_data]
	map = {i:j for i,j in zip(string.uppercase[1:col_end],range(year_start,year_end))}
	ws.cell(row=1,column=1).value = "Target Firm"
	for row in range(2,len(firms) + 2):
		firm = firms[row-2]
		ws.cell(row=row,column=1).value = firm
		for col in range(2,len(map) + 2): 
			cell = ws.cell(row=row,column=col)
			year = map[cell.column]
			ws.cell(row=1,column=col).value = year 
			if year in entry_data[firm][record]:
				cell.value = entry_data[firm][record][year]
	wb.save(output)
			
			
		
		# Cell = ws.cell(column=2,row=row)
		# company = unicode(Cell.value).strip()
		# if company and company != prevcompany and company in event_year_dict:
			# ws.cell(column=3,row=row).value = event_year_dict[company][0]
			# companies.append(company)
			# prevcompany = company
	# wb.save(target_firm_doc)	
	


def write_to_csv(output_file_name,entry_data):
	output_file = os.path.join(output_path,output_file_name)
	firms = [entry for entry in entry_data]
	sample = firms[0]
	with open(output_file,'wb') as csvfile:
		fieldnames = ["Target Firm"] + [field for field in entry_data[sample]]
		writer = csv.DictWriter(csvfile,fieldnames = fieldnames,extrasaction='ignore')
		writer.writeheader()
		for firm in firms:
			data = entry_data[firm]
			data.update({"Target Firm":firm})
			for record in data:
				if type(data[record]) == dict:
					data[record] = ', '.join("{!s}={!s}".format(*item) for item in data[record].iteritems())
			writer.writerow(data)

# doesnt work			
# def write_to_csv(output_file_name,entry_data):
	# output_file = os.path.join(output_path,output_file_name)
	# firms = [entry for entry in entry_data]
	# sample = firms[0]
	# fields = ["Target Firm"] + [field for field in entry_data[sample]]
	# with open(output_file,'wb') as csvfile:
		# writer = csv.DictWriter(csvfile,fieldnames = fields,extrasaction='ignore')
		# writer.writeheader()
		# for firm in firms:
			# data = entry_data[firm]
			# data.update({"Target Firm":firm})
			# for record in data:
				# if type(data[record]) == dict:
					# sub_fields = [field for field in data[record]]
					# sub_fields.insert(0,record)
					# print sub_fields
					# writer.fieldnames = sub_fields
					# writer.writeheader()
					# writer.writerow(data[record])
					# # data.pop(record)
					# print fields
					# fields = [field for field in fields if field != record]
			# writer.fieldnames = fields
			# writer.writerow(data)


# companies_sheet1 = get_company_names()
# companies_sheet2 = hash_event_years()
# for i in companies_sheet1:
# 	if i not in companies_sheet2:
# 		print i
# print len(get_company_names())
# print hash_event_years()
# print hash_year_values()
#write_event_years()
#create_new_xl()
# print check_pre_event_year()

if __name__ == "__main__":
	company_names = get_company_names()
	col_index_to_year_dict = hash_year_values()
	target_firms = build_target_firm_data(get_ranks(hash_event_years(company_names,col_index_to_year_dict),col_index_to_year_dict))
	# industry_firms = build_industry_groups(target_firms)
	# income_data = get_income_data(industry_firms,target_firms)
	# complete_data = get_match(income_data,target_firms)
	# create_new_xl("complete_data.xlsx", target_firms,["100_best_ranks","net_income_event_year_plus"])
	create_match_output("matches.xlsx",target_firms,omitted_records = ["100_best_ranks","net_income_event_year_plus"])
	create_year_output("One_hundred_best_ranks.xlsx",target_firms,"100_best_ranks")
	create_year_output("net_income.xlsx",target_firms,"net_income_event_year_plus")
